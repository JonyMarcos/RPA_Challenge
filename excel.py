import os
import datetime
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment


class ExcelWriter:
    def __init__(self, output_dir):
        self.output_dir = output_dir

    def get_output_filename(self):
        # Generate output filename based on current date and time.
        now = datetime.datetime.now()
        timestamp = now.strftime("%Y-%m-%d_%H-%M-%S")
        filename = f"gothamist_news_data_{timestamp}.xlsx"
        return filename

    def write_to_excel(self, data):
        # Write data to an Excel file.
        filename = self.get_output_filename()
        # Construct the full file path
        output_path = os.path.join(self.output_dir, filename)

        # Create a new Excel file
        wb = openpyxl.Workbook()
        sheet = wb.active

        # Define the headers
        headers = [
            "Search phrase",
            "Title",
            "Date",
            "Description",
            "Image URL",
            "Title Search Count",
            "Description Search Count",
            "Title Contains Money",
            "Description Contains Money",
        ]
        sheet.append(headers)

        # Add color to the header
        header_fill = PatternFill(
            start_color="00CCFF", end_color="00CCFF", fill_type="solid"
        )
        for cell in sheet[1]:
            cell.fill = header_fill

        # Write the data
        for news_item in data:
            sheet.append(news_item)

        # Adjust column widths and text alignment
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (
                max_length + 2
            ) * 1.2  # Adjusted width based on max length of data
            sheet.column_dimensions[column].width = adjusted_width
            for cell in col:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Set borders for all cells
        border = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
        )
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = border

        # Save the Excel file to the specified path
        wb.save(output_path)

        return output_path  # Return the saved file path
